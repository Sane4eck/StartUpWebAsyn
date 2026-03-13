from __future__ import annotations

import multiprocessing as mp
import queue
import threading
import time
from dataclasses import asdict
from pathlib import Path
from typing import Any, Callable

from openpyxl import load_workbook
from serial.tools import list_ports

from controller.runtime_types import PsuSnapshot, PsuTarget, VescSnapshot, VescTarget, make_command
from controller.workers import logger_worker_main, psu_worker_main, vesc_worker_main
from scheme.startup import StartupConfig


def _clamp(v: float, lo: float, hi: float) -> float:
    return max(lo, min(hi, float(v)))


def _interp_profile(xs: list[float], ys: list[float], x: float) -> float:
    if not xs or not ys:
        return 0.0
    if len(xs) == 1:
        return float(ys[0])
    if x <= xs[0]:
        return float(ys[0])
    if x >= xs[-1]:
        return float(ys[-1])

    for i in range(1, len(xs)):
        x0 = xs[i - 1]
        x1 = xs[i]
        if x <= x1:
            y0 = ys[i - 1]
            y1 = ys[i]
            if x1 == x0:
                return float(y1)
            a = (x - x0) / (x1 - x0)
            return float(y0 + a * (y1 - y0))
    return float(ys[-1])


def _load_profile_xlsx(path: str) -> tuple[list[float], list[float]]:
    wb = load_workbook(filename=path, data_only=True, read_only=True)
    ws = wb.active

    points: list[tuple[float, float]] = []
    for row in ws.iter_rows(values_only=True):
        if not row or len(row) < 2:
            continue
        try:
            t = float(row[0])
            rpm = float(row[1])
        except Exception:
            continue
        points.append((t, rpm))

    if not points:
        raise ValueError("Pump profile XLSX must contain at least two numeric columns: time, rpm")

    points.sort(key=lambda x: x[0])
    return [p[0] for p in points], [p[1] for p in points]


class _ProcHandle:
    def __init__(self, ctx: mp.context.BaseContext, name: str, target, args: tuple[Any, ...]):
        self.ctx = ctx
        self.name = name
        self.target = target
        self.args = args

        self.cmd_q = ctx.Queue()
        self.evt_q = ctx.Queue()
        self.stop_evt = ctx.Event()
        self.proc: mp.Process | None = None

    def start(self) -> None:
        if self.proc is not None and self.proc.is_alive():
            return
        self.stop_evt.clear()
        self.proc = self.ctx.Process(
            target=self.target,
            args=(self.name, self.cmd_q, self.evt_q, self.stop_evt, *self.args),
            name=f"{self.name}-proc",
            daemon=True,
        )
        self.proc.start()

    def stop(self) -> None:
        self.stop_evt.set()
        try:
            self.cmd_q.put_nowait({"kind": "__stop__", "payload": {}})
        except Exception:
            pass
        if self.proc is not None and self.proc.is_alive():
            self.proc.join(timeout=2.0)
        if self.proc is not None and self.proc.is_alive():
            self.proc.terminate()
            self.proc.join(timeout=1.0)

    def post(self, msg: dict[str, Any]) -> None:
        self.cmd_q.put(msg)

    def drain_events(self) -> list[dict[str, Any]]:
        out: list[dict[str, Any]] = []
        while True:
            try:
                out.append(self.evt_q.get_nowait())
            except queue.Empty:
                break
        return out


class WebControllerRuntime:
    def __init__(
        self,
        publish: Callable[[str, Any], None] | None = None,
        dt: float = 0.05,
    ):
        self.publish = publish or (lambda event, payload: None)

        self.dt = float(dt)
        self.ui_hz = 5.0
        self.log_hz = 5.0
        self.stale_worker_s = 1.0

        self._ui_dt = 1.0 / self.ui_hz
        self._log_dt = 1.0 / self.log_hz

        self._lock = threading.RLock()
        self._stop_evt = threading.Event()
        self._thread: threading.Thread | None = None

        self._ctx = mp.get_context("spawn")
        self._pump_proc = _ProcHandle(self._ctx, "pump", vesc_worker_main, (0.02, 0.01, 115200))
        self._starter_proc = _ProcHandle(self._ctx, "starter", vesc_worker_main, (0.02, 0.01, 115200))
        self._psu_proc = _ProcHandle(self._ctx, "psu", psu_worker_main, (0.02, 0.5, 0.2, 115200, 0.2, 1))
        self._logger_proc = _ProcHandle(self._ctx, "logger", logger_worker_main, ("logs",))

        self._cfg = StartupConfig()

        self._session_t0 = time.monotonic()
        self._stage_t0 = self._session_t0
        self._last_ui = 0.0
        self._last_log = 0.0

        self._stage = "idle"
        self._last_error = ""
        self._logger_path = ""

        self._pump_snap = VescSnapshot()
        self._starter_snap = VescSnapshot()
        self._psu_snap = PsuSnapshot()

        self._pump_manual = VescTarget(mode="rpm", value=0.0)
        self._starter_manual = VescTarget(mode="duty", value=0.0)
        self._psu_manual = PsuTarget(v=0.0, i=0.0, out=False)

        self._startup_active = False
        self._cooling_active = False
        self._cooling_until = 0.0

        self._pump_profile_active = False
        self._pump_profile_path = ""
        self._pump_profile_t: list[float] = []
        self._pump_profile_rpm: list[float] = []
        self._pump_profile_t0 = 0.0

        self._valve_macro_active = False
        self._valve_macro_t0 = 0.0

        self._holds: dict[str, float] = {}
        self._seq = 0

    def start(self) -> None:
        if self._thread is not None and self._thread.is_alive():
            return

        self._stop_evt.clear()
        self._pump_proc.start()
        self._starter_proc.start()
        self._psu_proc.start()
        self._logger_proc.start()

        self._thread = threading.Thread(
            target=self._run_loop,
            name="startup-web-runtime",
            daemon=True,
        )
        self._thread.start()

    def shutdown(self) -> None:
        self._stop_evt.set()

        if self._thread is not None and self._thread.is_alive():
            self._thread.join(timeout=2.0)

        self._safe_zero_outputs()

        self._pump_proc.stop()
        self._starter_proc.stop()
        self._psu_proc.stop()
        self._logger_proc.stop()

    def list_ports(self) -> list[str]:
        items = []
        for p in list_ports.comports():
            if p.device:
                items.append(str(p.device))
        return sorted(set(items))

    def snapshot(self) -> dict[str, Any]:
        return {
            "ports": self.list_ports(),
            "status": self._build_status(),
            "sample": self._build_sample(),
            "last_error": self._last_error,
        }

    def _next_seq(self) -> int:
        self._seq += 1
        return self._seq

    def _post(self, proc: _ProcHandle, worker: str, kind: str, payload: dict[str, Any]) -> None:
        proc.post(make_command(worker=worker, kind=kind, payload=payload, seq=self._next_seq()))

    def cmd_connect_pump(self, port: str) -> None:
        port = str(port or "").strip()
        if not port:
            raise ValueError("Pump COM port is empty")
        self._post(self._pump_proc, "pump", "connect", {"port": port})

    def cmd_disconnect_pump(self) -> None:
        self._post(self._pump_proc, "pump", "disconnect", {})

    def cmd_connect_starter(self, port: str) -> None:
        port = str(port or "").strip()
        if not port:
            raise ValueError("Starter COM port is empty")
        self._post(self._starter_proc, "starter", "connect", {"port": port})

    def cmd_disconnect_starter(self) -> None:
        self._post(self._starter_proc, "starter", "disconnect", {})

    def cmd_connect_psu(self, port: str) -> None:
        port = str(port or "").strip()
        if not port:
            raise ValueError("PSU COM port is empty")
        self._post(self._psu_proc, "psu", "connect", {"port": port})

    def cmd_disconnect_psu(self) -> None:
        self._post(self._psu_proc, "psu", "disconnect", {})

    def cmd_set_pole_pairs_pump(self, pole_pairs: int) -> None:
        self._post(self._pump_proc, "pump", "set_pole_pairs", {"pole_pairs": max(1, int(pole_pairs))})

    def cmd_set_pole_pairs_starter(self, pole_pairs: int) -> None:
        self._post(self._starter_proc, "starter", "set_pole_pairs", {"pole_pairs": max(1, int(pole_pairs))})

    def cmd_ready(self, prefix: str = "manual") -> None:
        now = time.monotonic()
        with self._lock:
            self._cancel_automation_locked()
            self._pump_manual = VescTarget(mode="rpm", value=0.0)
            self._starter_manual = VescTarget(mode="duty", value=0.0)
            self._psu_manual = PsuTarget(v=0.0, i=0.0, out=False)
            self._stage = "ready"
            self._stage_t0 = now
            self._session_t0 = now
            self._last_error = ""
        self._post(self._logger_proc, "logger", "open", {"prefix": prefix or "manual"})
        self._publish("status", {**self._build_status(), "ready": True})

    def cmd_update_reset(self) -> None:
        self._publish("status", {**self._build_status(), "reset_plot": True})

    def cmd_run_cycle(self) -> None:
        now = time.monotonic()
        with self._lock:
            self._cancel_automation_locked()
            self._startup_active = True
            self._stage = "starter"
            self._stage_t0 = now
            self._holds.clear()

    def cmd_cooling_cycle(self, value: float) -> None:
        duration_s = max(0.1, float(value))
        now = time.monotonic()
        with self._lock:
            self._cancel_automation_locked()
            self._cooling_active = True
            self._cooling_until = now + duration_s
            self._stage = "cooling"
            self._stage_t0 = now

    def cmd_stop_all(self) -> None:
        now = time.monotonic()
        with self._lock:
            self._cancel_automation_locked()
            self._pump_manual = VescTarget(mode="rpm", value=0.0)
            self._starter_manual = VescTarget(mode="duty", value=0.0)
            self._psu_manual = PsuTarget(v=self._psu_manual.v, i=self._psu_manual.i, out=False)
            self._stage = "stop"
            self._stage_t0 = now

    def cmd_valve_on(self) -> None:
        now = time.monotonic()
        with self._lock:
            self._cancel_startup_only_locked()
            self._valve_macro_active = True
            self._valve_macro_t0 = now
            if self._stage not in {"ready", "idle", "stop"}:
                self._stage = "manual"
                self._stage_t0 = now

    def cmd_valve_off(self) -> None:
        now = time.monotonic()
        with self._lock:
            self._valve_macro_active = False
            self._psu_manual.out = False
            if self._stage not in {"ready", "idle", "stop"}:
                self._stage = "manual"
                self._stage_t0 = now

    def cmd_start_pump_profile(self, path: str) -> None:
        raw = str(path or "").strip()
        if not raw:
            raise ValueError("Pump profile path is empty")
        if not Path(raw).exists():
            raise ValueError(f"Pump profile file not found: {raw}")

        xs, ys = _load_profile_xlsx(raw)
        now = time.monotonic()

        with self._lock:
            self._cancel_automation_locked()
            self._pump_profile_active = True
            self._pump_profile_path = raw
            self._pump_profile_t = xs
            self._pump_profile_rpm = ys
            self._pump_profile_t0 = now
            self._stage = "pump_profile"
            self._stage_t0 = now

    def cmd_stop_pump_profile(self) -> None:
        now = time.monotonic()
        with self._lock:
            self._pump_profile_active = False
            self._pump_profile_path = ""
            self._pump_profile_t = []
            self._pump_profile_rpm = []
            if self._stage == "pump_profile":
                self._stage = "manual"
                self._stage_t0 = now

    def cmd_set_pump_rpm(self, value: float) -> None:
        now = time.monotonic()
        with self._lock:
            self._cancel_automation_locked()
            self._pump_manual = VescTarget(mode="rpm", value=float(value))
            if self._stage not in {"idle", "ready", "stop"}:
                self._stage = "manual"
                self._stage_t0 = now

    def cmd_set_pump_duty(self, value: float) -> None:
        now = time.monotonic()
        with self._lock:
            self._cancel_automation_locked()
            self._pump_manual = VescTarget(mode="duty", value=_clamp(value, 0.0, 1.0))
            if self._stage not in {"idle", "ready", "stop"}:
                self._stage = "manual"
                self._stage_t0 = now

    def cmd_set_starter_rpm(self, value: float) -> None:
        now = time.monotonic()
        with self._lock:
            self._cancel_automation_locked()
            self._starter_manual = VescTarget(mode="rpm", value=float(value))
            if self._stage not in {"idle", "ready", "stop"}:
                self._stage = "manual"
                self._stage_t0 = now

    def cmd_set_starter_duty(self, value: float) -> None:
        now = time.monotonic()
        with self._lock:
            self._cancel_automation_locked()
            self._starter_manual = VescTarget(mode="duty", value=_clamp(value, 0.0, 1.0))
            if self._stage not in {"idle", "ready", "stop"}:
                self._stage = "manual"
                self._stage_t0 = now

    def cmd_psu_set_vi(self, v: float, i: float) -> None:
        now = time.monotonic()
        with self._lock:
            self._cancel_automation_locked()
            self._psu_manual.v = float(v)
            self._psu_manual.i = float(i)
            if self._stage not in {"idle", "ready", "stop"}:
                self._stage = "manual"
                self._stage_t0 = now

    def cmd_psu_output(self, value: bool) -> None:
        now = time.monotonic()
        with self._lock:
            self._cancel_automation_locked()
            self._psu_manual.out = bool(value)
            if self._stage not in {"idle", "ready", "stop"}:
                self._stage = "manual"
                self._stage_t0 = now

    def _publish(self, event: str, payload: Any) -> None:
        try:
            self.publish(event, payload)
        except Exception:
            pass

    def _run_loop(self) -> None:
        next_tick = time.monotonic()

        while not self._stop_evt.is_set():
            now = time.monotonic()
            if now < next_tick:
                time.sleep(min(0.005, next_tick - now))
                continue

            self._tick(now)
            next_tick += self.dt

            if next_tick < now - self.dt:
                next_tick = now + self.dt

    def _tick(self, now: float) -> None:
        self._drain_worker_events()

        with self._lock:
            pump_target = VescTarget(self._pump_manual.mode, self._pump_manual.value)
            starter_target = VescTarget(self._starter_manual.mode, self._starter_manual.value)
            psu_target = PsuTarget(self._psu_manual.v, self._psu_manual.i, self._psu_manual.out)

            if self._pump_profile_active:
                dt_profile = now - self._pump_profile_t0
                pump_target = VescTarget(
                    mode="rpm",
                    value=_interp_profile(self._pump_profile_t, self._pump_profile_rpm, dt_profile),
                )
                if dt_profile >= (self._pump_profile_t[-1] if self._pump_profile_t else 0.0):
                    self._pump_profile_active = False
                    if self._stage == "pump_profile":
                        self._stage = "manual"
                        self._stage_t0 = now

            if self._startup_active:
                self._tick_startup_locked(now)
                pump_target = VescTarget(self._pump_manual.mode, self._pump_manual.value)
                starter_target = VescTarget(self._starter_manual.mode, self._starter_manual.value)

            if self._cooling_active:
                if now >= self._cooling_until:
                    self._cooling_active = False
                    self._stage = "stop"
                    self._stage_t0 = now
                else:
                    starter_target = VescTarget(mode="duty", value=0.0)
                    psu_target = PsuTarget(v=psu_target.v, i=psu_target.i, out=False)

            if self._valve_macro_active:
                dt_valve = now - self._valve_macro_t0
                if dt_valve < self._cfg.valve_boost_s:
                    psu_target = PsuTarget(
                        v=self._cfg.valve_boost_v,
                        i=self._cfg.valve_boost_i,
                        out=True,
                    )
                else:
                    psu_target = PsuTarget(
                        v=self._cfg.valve_hold_v,
                        i=self._cfg.valve_hold_i,
                        out=True,
                    )

        self._post(self._pump_proc, "pump", "set_target", {"mode": pump_target.mode, "value": pump_target.value})
        self._post(self._starter_proc, "starter", "set_target", {"mode": starter_target.mode, "value": starter_target.value})
        self._post(self._psu_proc, "psu", "set_target", {"v": psu_target.v, "i": psu_target.i, "out": psu_target.out})

        self._check_stale(now)

        if (now - self._last_ui) >= self._ui_dt:
            self._last_ui = now
            self._publish("status", self._build_status())
            self._publish("sample", self._build_sample())

        if (now - self._last_log) >= self._log_dt:
            self._last_log = now
            self._post(self._logger_proc, "logger", "row", self._build_log_row())

    def _drain_worker_events(self) -> None:
        for proc in (self._pump_proc, self._starter_proc, self._psu_proc, self._logger_proc):
            for evt in proc.drain_events():
                worker = str(evt.get("worker", ""))
                kind = str(evt.get("kind", ""))
                payload = evt.get("payload", {}) or {}

                if worker == "pump" and kind == "snapshot":
                    self._pump_snap = VescSnapshot(**payload)
                elif worker == "starter" and kind == "snapshot":
                    self._starter_snap = VescSnapshot(**payload)
                elif worker == "psu" and kind == "snapshot":
                    self._psu_snap = PsuSnapshot(**payload)
                elif worker == "logger" and kind == "state":
                    self._logger_path = str(payload.get("log_path", "") or "")
                    err = payload.get("error")
                    if err:
                        self._last_error = str(err)
                        self._publish("error", self._last_error)

                if kind == "error":
                    msg = str(payload.get("message", "worker error"))
                    self._last_error = f"{worker}: {msg}"
                    self._publish("error", self._last_error)

    def _check_stale(self, now: float) -> None:
        snaps = [self._pump_snap, self._starter_snap]
        for snap in snaps:
            if snap.connected and (now - snap.ts) > self.stale_worker_s:
                self._fault_locked(now, "Worker stale timeout")
                return

    def _tick_startup_locked(self, now: float) -> None:
        cfg = self._cfg
        stage = self._stage
        stage_dt = now - self._stage_t0

        starter_rpm = float(self._starter_snap.rpm_mech)
        pump_rpm = float(self._pump_snap.rpm_mech)

        if stage == "starter":
            duty = self._starter_step_duty(starter_rpm)
            self._starter_manual = VescTarget(mode="duty", value=duty)

            if self._hold_true_locked(
                "to_fuelramp",
                starter_rpm >= cfg.to_fuelramp_starter_rpm,
                now,
                cfg.to_fuelramp_hold_s,
            ):
                self._stage = "fuelramp"
                self._stage_t0 = now
                self._holds.clear()
                self._valve_macro_active = True
                self._valve_macro_t0 = now
                return

            if stage_dt >= cfg.starter_timeout_s:
                self._fault_locked(now, "Starter timeout")
                return

        elif stage == "fuelramp":
            duty = self._starter_step_duty(starter_rpm)
            self._starter_manual = VescTarget(mode="duty", value=duty)

            if self._hold_true_locked(
                "valve_close",
                starter_rpm >= cfg.valve_close_rpm,
                now,
                cfg.valve_close_hold_s,
            ):
                self._valve_macro_active = False

            if self._hold_true_locked(
                "starter_off",
                starter_rpm >= cfg.starter_off_rpm,
                now,
                cfg.starter_off_hold_s,
            ):
                self._starter_manual = VescTarget(mode="duty", value=0.0)

            if self._hold_true_locked(
                "to_running",
                starter_rpm >= cfg.to_running_starter_rpm,
                now,
                cfg.to_running_hold_s,
            ):
                self._startup_active = False
                self._stage = "running"
                self._stage_t0 = now
                self._holds.clear()
                self._valve_macro_active = False
                self._starter_manual = VescTarget(mode="duty", value=0.0)
                self._pump_manual = VescTarget(mode="rpm", value=pump_rpm)
                return

            if stage_dt >= cfg.fuelramp_timeout_s:
                self._fault_locked(now, "FuelRamp timeout")
                return

    def _starter_step_duty(self, starter_rpm: float) -> float:
        duty = float(self._cfg.starter_steps[0][1])
        for rpm_thr, d in self._cfg.starter_steps:
            if starter_rpm >= float(rpm_thr):
                duty = float(d)
            else:
                break
        return duty

    def _hold_true_locked(self, key: str, cond: bool, now: float, hold_s: float) -> bool:
        if cond:
            self._holds.setdefault(key, now)
            return (now - self._holds[key]) >= float(hold_s)
        self._holds.pop(key, None)
        return False

    def _fault_locked(self, now: float, reason: str) -> None:
        self._cancel_automation_locked()
        self._pump_manual = VescTarget(mode="rpm", value=0.0)
        self._starter_manual = VescTarget(mode="duty", value=0.0)
        self._psu_manual.out = False
        self._stage = "fault"
        self._stage_t0 = now
        self._last_error = str(reason)
        self._publish("error", self._last_error)

    def _cancel_startup_only_locked(self) -> None:
        self._startup_active = False
        self._holds.clear()

    def _cancel_automation_locked(self) -> None:
        self._startup_active = False
        self._cooling_active = False
        self._cooling_until = 0.0

        self._pump_profile_active = False
        self._pump_profile_path = ""
        self._pump_profile_t = []
        self._pump_profile_rpm = []
        self._pump_profile_t0 = 0.0

        self._valve_macro_active = False
        self._valve_macro_t0 = 0.0

        self._holds.clear()

    def _safe_zero_outputs(self) -> None:
        try:
            self._post(self._pump_proc, "pump", "set_target", {"mode": "rpm", "value": 0.0})
            self._post(self._starter_proc, "starter", "set_target", {"mode": "duty", "value": 0.0})
            self._post(self._psu_proc, "psu", "set_target", {"v": self._psu_manual.v, "i": self._psu_manual.i, "out": False})
            time.sleep(0.1)
        except Exception:
            pass

    def _build_status(self) -> dict[str, Any]:
        return {
            "stage": self._stage,
            "connected": {
                "pump": self._pump_snap.connected,
                "starter": self._starter_snap.connected,
                "psu": self._psu_snap.connected,
            },
            "log_path": self._logger_path,
            "pump_profile": {
                "active": self._pump_profile_active,
                "path": self._pump_profile_path,
            },
            "valve_macro": {
                "active": self._valve_macro_active,
            },
        }

    def _build_sample(self) -> dict[str, Any]:
        return {
            "t": max(0.0, time.monotonic() - self._session_t0),
            "stage": self._stage,
            "connected": {
                "pump": self._pump_snap.connected,
                "starter": self._starter_snap.connected,
                "psu": self._psu_snap.connected,
            },
            "pump": asdict(self._pump_snap),
            "starter": asdict(self._starter_snap),
            "psu": asdict(self._psu_snap),
        }

    def _build_log_row(self) -> dict[str, Any]:
        sample = self._build_sample()
        pump = sample["pump"]
        starter = sample["starter"]
        psu = sample["psu"]

        return {
            "t": round(float(sample["t"]), 6),
            "stage": sample["stage"],
            "pump_rpm": pump.get("rpm_mech", 0.0),
            "pump_duty": pump.get("duty", 0.0),
            "pump_current": pump.get("current_motor", 0.0),
            "starter_rpm": starter.get("rpm_mech", 0.0),
            "starter_duty": starter.get("duty", 0.0),
            "starter_current": starter.get("current_motor", 0.0),
            "psu_v_set": psu.get("v_set", 0.0),
            "psu_i_set": psu.get("i_set", 0.0),
            "psu_v_out": psu.get("v_out", 0.0),
            "psu_i_out": psu.get("i_out", 0.0),
            "psu_output": int(bool(psu.get("output", False))),
        }
